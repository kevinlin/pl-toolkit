import argparse
import pprint
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook


def extract_user_row_mappings(df):
    """Extract team members and their corresponding row indices from the timesheet."""
    user_row_index = df[df.iloc[:, 0] == 'User'].index
    summary_categories = ["Absenses [h]", "Networking and administration, personal development", "Operational hours"]

    if not user_row_index.empty:
        # Extract names from the next rows until an empty row is encountered
        start_index = user_row_index[0] + 1
        user_rows = {}
        category_row_indices = {}  # Store category row indices for each user
        cur_name = None

        for row_idx in range(start_index, len(df)):
            name = str(df.iloc[row_idx, 0]).strip()
            if name and name.lower() != 'nan' and not pd.isna(name):
                cur_name = name
                user_rows[name] = row_idx
                category_row_indices[name] = {}
            else:
                # Identify category row indices
                category = str(df.iloc[row_idx, 11]).strip()  # Col L
                if category in summary_categories:
                    category_row_indices[cur_name][category] = row_idx

                continue

        return user_rows, category_row_indices

    return {}, {}


def extract_date_col_mappings(df):
    """Extract valid workdays from the timesheet, excluding weekends and limiting to last working Friday."""
    columns = df.iloc[2, 15:].dropna().astype(str).tolist()  # Start at Col P (index 15) and go to Col AT (index 45)
    valid_days = {}

    today = datetime.now()
    previous_month = False
    if today.day <= 5:
        # Take last day of previous month if today is the first week of the month
        previous_month = True
        last_day = today.replace(day=1) - timedelta(days=1)
        today = last_day

    for idx, col in enumerate(columns):  # Track column index
        try:
            day_str, weekday = col.split(", ")  # Extract numeric day and weekday
            day = int(day_str)  # Convert day to integer
            date = today.replace(day=day)
            if weekday not in ["Sa", "Su"]:  # Only include weekdays up to yesterday
                if (previous_month and day <= last_day.day) or (not previous_month and day < today.day):
                    valid_days[date.strftime("%a, %b-%d")] = idx + 15
        except ValueError:
            continue

    return valid_days


def read_timesheet_entries_by_users(df, user_row_mappings, date_col_mappings):
    """Read timesheet entries for each user by valid workdays."""
    timesheet_data = pd.DataFrame(index=user_row_mappings.keys(), columns=date_col_mappings.keys())

    for user, row_idx in user_row_mappings.items():
        for day_str, col_idx in date_col_mappings.items():
            target_hours = df.iloc[row_idx + 2, col_idx]
            target_hours = int(target_hours) if pd.notna(target_hours) else 0
            actual_hours = df.iloc[row_idx + 3, col_idx]
            actual_hours = int(actual_hours) if pd.notna(actual_hours) else 0
            timesheet_data.at[user, day_str] = target_hours - actual_hours
        timesheet_data.at[user, "Submitted?"] = df.iloc[row_idx, 2] == 1

    return timesheet_data


def summarise_time_distribution(df, category_row_indices, date_col_mappings):
    """Summarise the hours each user spent separately on predefined categories."""
    summary_data = {
        user: {"Absenses [h]": 0, "Networking and administration, personal development": 0, "Operational hours": 0} for
        user in category_row_indices}

    for user, categories in category_row_indices.items():
        for category, row_idx in categories.items():
            # print(f"User: {user}, Category: {category}->")
            # for _, col_idx in date_col_mappings.items():
            #     print(f"[{row_idx+1}:{col_idx}]: {str(df.iloc[row_idx+1, col_idx])}")

            cell_values = df.iloc[row_idx + 1, [col_idx for _, col_idx in date_col_mappings.items()]]
            total_hours = cell_values[cell_values.apply(lambda x: isinstance(x, (int, float)) and pd.notna(x))].sum()
            summary_data[user][category] = total_hours

    return pd.DataFrame.from_dict(summary_data, orient='index')


def main():
    parser = argparse.ArgumentParser(description='Process Vertec timesheet')
    parser.add_argument('file_path', type=str, help='Path to the input Excel file')
    file_path = parser.parse_args().file_path

    # Load and process the timesheet
    df = pd.read_excel(file_path, sheet_name="Sheet2", skiprows=0)

    pp = pprint.PrettyPrinter(indent=4)

    # Extract data
    user_row_mappings, category_row_indices = extract_user_row_mappings(df)
    # pp.pprint(user_row_mappings)
    pp.pprint(category_row_indices)
    date_col_mappings = extract_date_col_mappings(df)
    # print(f"Days to check: {date_col_mappings}")
    timesheet_entries = read_timesheet_entries_by_users(df, user_row_mappings, date_col_mappings)
    print(timesheet_entries)
    # Load workbook with data_only=True to get cell values (not formulas)
    worksheet = load_workbook(filename=file_path, data_only=True)["Sheet2"]  # Replace with your sheet name
    dfs = pd.DataFrame(list(worksheet.iter_rows(values_only=True)))
    summary_data = summarise_time_distribution(dfs, category_row_indices, date_col_mappings)
    print(summary_data)

    timesheet_entries.to_csv("output/timesheet_entries.csv", index=True)
    summary_data.to_csv("output/time_distribution.csv", index=True)


if __name__ == "__main__":
    main()
